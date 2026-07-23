"""
Per-strategy trade blotter + portfolio Excel aggregator (src/strategies/utils/blotter.py).
Validates the blotter math (traded $ = Δposition, fee/spread from bps, side, first-day establish,
zero-trade rows dropped) and that the workbook has one sheet per sleeve + a summary.
"""
from __future__ import annotations

import numpy as np
import pandas as pd
import openpyxl

from src.strategies.utils.blotter import trade_blotter, write_trades_excel


def test_trade_blotter_math():
    idx = pd.bdate_range("2024-01-01", periods=3)
    # AAA: 0.10 -> 0.10 -> 0.05 ; BBB: -0.20 -> -0.20 -> -0.20 (BBB never trades after day 1)
    w = pd.DataFrame({"AAA": [0.10, 0.10, 0.05], "BBB": [-0.20, -0.20, -0.20]}, index=idx)
    bl = trade_blotter(w, capital=1_000_000, fee_bps=2.0, spread_bps=8.0, sleeve="ls_equity")

    # day1 establishes both (AAA +100k BUY, BBB -200k SELL); day3 AAA sells 50k; BBB day2/3 no trade
    d1 = bl[bl["date"] == idx[0]].set_index("instrument")
    assert d1.loc["AAA", "traded_usd"] == 100_000 and d1.loc["AAA", "side"] == "BUY"
    assert d1.loc["BBB", "traded_usd"] == -200_000 and d1.loc["BBB", "side"] == "SELL"
    assert d1.loc["AAA", "fee_usd"] == 100_000 * 2e-4 and d1.loc["AAA", "spread_usd"] == 100_000 * 8e-4
    assert d1.loc["AAA", "cost_usd"] == 100_000 * 10e-4
    d3 = bl[bl["date"] == idx[2]].set_index("instrument")
    assert d3.loc["AAA", "traded_usd"] == -50_000 and d3.loc["AAA", "side"] == "SELL"
    assert "BBB" not in d3.index                        # BBB didn't trade on day 3 -> no row
    assert bl[bl["date"] == idx[1]].empty                # day2: both names flat -> NO trade rows

    print("\n=== SANITY CHECK: trade blotter math ===")
    print(f"  rows={len(bl)}; day1 AAA BUY $100k (fee ${d1.loc['AAA','fee_usd']:.0f}, "
          f"spread ${d1.loc['AAA','spread_usd']:.0f}); day3 AAA SELL $50k; zero-trades dropped. Validated.")


def test_shares_and_transition():
    idx = pd.bdate_range("2024-01-01", periods=3)
    # AAA: held (w=0.10) then dropped day3 ; BBB: new only on day3 (w=0.20). Prices drift.
    w = pd.DataFrame({"AAA": [0.10, 0.10, 0.00], "BBB": [0.00, 0.00, 0.20]}, index=idx)
    px = pd.DataFrame({"AAA": [10.0, 12.5, 12.0], "BBB": [5.0, 5.0, 5.0]}, index=idx)
    bl = trade_blotter(w, capital=1_000_000, fee_bps=2.0, spread_bps=8.0, sleeve="ls_equity", prices=px)
    g = bl.set_index(["date", "instrument"])

    # day1: BUY 10,000 shares of AAA ($100k / $10)
    assert g.loc[(idx[0], "AAA"), "shares_traded"] == 10_000 and g.loc[(idx[0], "AAA"), "side"] == "BUY"
    # day2: $ position unchanged ($100k) but price rose 10->12.5 => hold 8,000 shares => SELL 2,000
    assert g.loc[(idx[1], "AAA"), "shares_traded"] == -2_000        # price-drift share rebalance
    assert g.loc[(idx[1], "AAA"), "traded_usd"] == -25_000
    # day3: AAA fully EXITED (sell the 8,000 held) AND BBB NEWLY BOUGHT (40,000) -> two moves, two fees
    assert g.loc[(idx[2], "AAA"), "shares_traded"] == -8_000 and g.loc[(idx[2], "AAA"), "side"] == "SELL"
    assert g.loc[(idx[2], "BBB"), "shares_traded"] == 40_000 and g.loc[(idx[2], "BBB"), "side"] == "BUY"
    assert g.loc[(idx[2], "AAA"), "fee_usd"] == abs(-8_000 * 12.0) * 2e-4    # fee per move on shares·price

    print("\n=== SANITY CHECK: share-accurate blotter + transitions ===")
    print(f"  day1 BUY 10,000 AAA; day2 SELL 2,000 AAA (price drift 10->12.5 at constant $100k); "
          f"day3 SELL 8,000 AAA (exit) + BUY 40,000 BBB (new) — each a separate move+fee. Validated.")


def test_write_trades_excel(tmp_path):
    idx = pd.bdate_range("2024-01-01", periods=3)
    w = pd.DataFrame({"AAA": [0.1, 0.2, 0.1]}, index=idx)
    trades = {"ls_equity": trade_blotter(w, 1e6, 2.0, 8.0, "ls_equity"),
              "trend_cta": trade_blotter(w * -1, 1e6, 2.0, 8.0, "trend_cta"),
              "long_book": None}                          # a sleeve with no trades still gets a sheet
    path = tmp_path / "trades.xlsx"
    write_trades_excel(trades, path)
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == "summary"
    assert set(["ls_equity", "trend_cta", "long_book"]).issubset(set(wb.sheetnames))
    summ = pd.read_excel(path, sheet_name="summary").set_index("sleeve")
    assert summ.loc["ls_equity", "n_trades"] > 0 and summ.loc["ls_equity", "total_cost_usd"] > 0
    print("\n=== SANITY CHECK: portfolio trade workbook ===")
    print(f"  sheets={wb.sheetnames}; ls_equity total cost "
          f"${summ.loc['ls_equity','total_cost_usd']:.0f}. Validated.")


if __name__ == "__main__":
    test_trade_blotter_math()
